"""
app.py — Dienstplanning solver-service (FastAPI + OR-Tools CP-SAT)

Endpoint:
  POST /solve-weekday   → berekent de doordeweekse dienstplanning
  GET  /health          → healthcheck

De service is bewust dun: hij vertaalt een verzoek naar de generieke
RosterEngine en geeft het resultaat per datum terug. Dezelfde engine
bedient later ook de dagplanning (ander endpoint, andere slots).

CORS staat open zodat de React-site (localhost of Netlify) hem mag aanroepen.
"""

from datetime import date, timedelta, datetime as dt
from io import BytesIO
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import openpyxl

from roster_engine import RosterEngine, Staff, Slot, Period, Avail, SoftWeights

SKILL = "DIENST_WEEKDAG"
WD_CODES = ["ma", "di", "wo", "do", "vr", "za", "zo"]

app = FastAPI(title="Dienstplanning solver", version="1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"],
)


class DoctorIn(BaseModel):
    id: str
    carryIn: int = 0
    fixedOff: list[str] = []          # weekdagcodes, bv. ["wo"]
    preferOff: list[str] = []         # weekdagcodes, zacht
    biweeklyOff: list[dict] = []      # [{"day":"wo","parity":"even"|"oneven"}] — hard, ISO-weekpariteit
    absences: list[dict] = []         # [{"from":"2026-07-06","to":"2026-07-10","type":"vakantie"}]


class SolveReq(BaseModel):
    start: str                        # ISO-datum; wordt naar maandag uitgelijnd
    weeks: int = 4
    doctors: list[DoctorIn]
    weights: dict | None = None


@app.get("/health")
def health():
    return {"ok": True}


# ---------------------------------------------------------------------------
# /parse-requirements
# Accepteert OK- en/of PBK-Excel (multipart), retourneert geldige PLA-slots.
# Een PLA is geldig als: niet doorgestreept EN niet overgedragen ("PLA > X").
# Structuur retour: { requirements: { "2026-01-05": { "AM": ["OK"], "PM": [] } } }
# ---------------------------------------------------------------------------

import re as _re
_PLA_RE  = _re.compile(r'^PLA(\s|>|\(|-|$)', _re.IGNORECASE)
_DAYS_NL = {"MAANDAG", "DINSDAG", "WOENSDAG", "DONDERDAG", "VRIJDAG"}
# Per dag-blok: (datum_kol, ochtend_kol, middag_kol) voor weken 1-5
# Sommige maanden hebben 5 voorkomens van een weekdag (bv. juni heeft 5 woensdagen)
_WEEK_COLS = [(2, 2, 3), (4, 4, 5), (6, 6, 7), (8, 8, 9), (10, 10, 11)]


def _extract(content: bytes, kind: str, result: dict, min_date: str | None = None):
    """Parseer één Excel-bestand en voeg geldige PLA-requirements toe aan result."""
    wb_v = openpyxl.load_workbook(BytesIO(content), data_only=True)
    wb_f = openpyxl.load_workbook(BytesIO(content), data_only=False)

    for sheet in wb_v.sheetnames:
        wsv = wb_v[sheet]
        wsf = wb_f[sheet]

        for r in range(1, wsv.max_row + 1):
            a = wsv.cell(r, 1).value
            if not (a and str(a).strip().upper() in _DAYS_NL):
                continue

            # Verzamel rijen onder dit dag-anker tot de volgende dag
            block = []
            rr = r + 1
            while rr <= wsv.max_row:
                nxt = wsv.cell(rr, 1).value
                if nxt and str(nxt).strip().upper() in _DAYS_NL:
                    break
                block.append(rr)
                rr += 1

            for (dc, oc, mc) in _WEEK_COLS:
                dval = wsv.cell(r, dc).value
                if not isinstance(dval, dt):
                    continue
                date_str = dval.strftime("%Y-%m-%d")
                if min_date and date_str < min_date:
                    continue   # verleden week — negeren

                for (col, period) in [(oc, "AM"), (mc, "PM")]:
                    for br in block:
                        v = wsv.cell(br, col).value
                        if not (v and _PLA_RE.match(str(v).strip())):
                            continue
                        cell_f = wsf.cell(br, col)
                        struck = bool(cell_f.font and cell_f.font.strikethrough)
                        transfer = ">" in str(v)   # "PLA > NCH" = overgedragen
                        if struck or transfer:
                            continue
                        # Geldige requirement — voeg toe (dedupliceer)
                        slot = result.setdefault(date_str, {"AM": [], "PM": []})
                        if kind not in slot[period]:
                            slot[period].append(kind)


class DagDoctorIn(BaseModel):
    id: str
    role:        str = "dokter"
    activityIds: list[str] = []
    fixedOff:    list[str] = []
    preferOff:   list[str] = []
    biweeklyOff: list[dict] = []      # [{"day":"wo","parity":"even"|"oneven"}]
    manualOff:   list[dict] = []      # [{"date":"2026-07-06","period":"AM"}] — VRIJ-cel, hard per dagdeel
    absences:    list[dict] = []


def _is_biweekly_off(rules: list[dict], date_str: str, wd: str) -> bool:
    """'Om de week vrij' — hard, alleen op even/oneven ISO-weken (isocalendar,
    zelfde algoritme als de frontend's isoWeek())."""
    if not rules:
        return False
    parity = "even" if date.fromisoformat(date_str).isocalendar()[1] % 2 == 0 else "oneven"
    return any(r.get("day") == wd and r.get("parity") == parity for r in rules)


class DagplanningReqBody(BaseModel):
    requirements: dict                              # { "2026-01-05": { "AM": ["OK"], "PM": ["PBK"] } }
    doctors:      list[DagDoctorIn]
    reqActMap:    dict = {"OK": "act_ok", "PBK": "act_pbk", "Poli": "act_poli"}
    priorTotals:  dict = {}                         # { staffId: aantal reeds geplande slots dit jaar } (regel 3)


def _doctor_available(doc: DagDoctorIn, date_str: str, period: str | None = None) -> bool:
    """Een arts is beschikbaar op een dagdeel als hij niet vrij is en niet afwezig."""
    wd = WD_CODES[date.fromisoformat(date_str).weekday()]
    if wd in doc.fixedOff:
        return False
    if _is_biweekly_off(doc.biweeklyOff, date_str, wd):
        return False
    for mo in doc.manualOff:                      # handmatige/auto VRIJ-cel = hard vrij
        if mo.get("date") == date_str and (period is None or mo.get("period") == period):
            return False
    for ab in doc.absences:
        f, t = ab.get("from"), ab.get("to")
        if f and t and f <= date_str <= t:
            return False
    return True


@app.post("/solve-dagplanning")
def solve_dagplanning(req: DagplanningReqBody):
    """
    Volledige arts-planningsketen per dagdeel:
      beschikbaar = #artsen(rol=dokter) - #vrij - #scholing
      1. Vul OK + PBK uit beschikbaar. Bij tekort: PBK eerst teruggeven, dan OK.
      2. Poli = beschikbaar - (OK + PBK die overeind blijven).
      3. CP-SAT verdeelt de overeind gebleven OK/PBK/Poli eerlijk over de artsen.
    Retourneert toewijzingen + teruggeef-lijst (returned) + poli-vereisten.
    """
    doctors    = [d for d in req.doctors if d.role in ("dokter", "arts")]
    ok_act     = req.reqActMap.get("OK",   "act_ok")
    pbk_act    = req.reqActMap.get("PBK",  "act_pbk")
    poli_act   = req.reqActMap.get("Poli", "act_poli")

    monday_idx: dict[date, int] = {}
    slots: list[Slot] = []
    returned: list[dict] = []          # teruggegeven OK/PBK door tekort
    poli_added: dict[str, dict] = {}   # { date: { AM: n, PM: n } } voor terugkoppeling UI

    for date_str in sorted(req.requirements):
        d = date.fromisoformat(date_str)
        monday = d - timedelta(days=d.weekday())
        if monday not in monday_idx:
            monday_idx[monday] = len(monday_idx)
        wk  = monday_idx[monday]
        seq = d.weekday()

        for period_str in ["AM", "PM"]:
            period_enum = Period.AM if period_str == "AM" else Period.PM
            reqs = req.requirements[date_str].get(period_str, [])
            n_ok  = reqs.count("OK")
            n_pbk = reqs.count("PBK")

            # beschikbare artsen dit dagdeel
            avail_n = sum(1 for doc in doctors if _doctor_available(doc, date_str, period_str))

            # 1. OK + PBK passend maken op beschikbaar; PBK eerst teruggeven
            keep_ok, keep_pbk = n_ok, n_pbk
            while keep_ok + keep_pbk > avail_n:
                if keep_pbk > 0:
                    keep_pbk -= 1
                    returned.append({"date": date_str, "period": period_str, "type": "PBK"})
                elif keep_ok > 0:
                    keep_ok -= 1
                    returned.append({"date": date_str, "period": period_str, "type": "OK"})
                else:
                    break

            # 2. Poli = rest van de beschikbare artsen
            n_poli = max(0, avail_n - keep_ok - keep_pbk)
            if n_poli > 0:
                poli_added.setdefault(date_str, {})[period_str] = n_poli

            # 3. slots bouwen (index per type voor unieke id's)
            def _add(act_id, label, count):
                for i in range(count):
                    slots.append(Slot(
                        id=f"{date_str}__{period_str}__{label}__{i}",
                        date=date_str, period=period_enum,
                        required_skill=act_id, demand=1,
                        week=wk, seq=wk * 10 + seq,
                    ))
            _add(ok_act,   "OK",   keep_ok)
            _add(pbk_act,  "PBK",  keep_pbk)
            _add(poli_act, "Poli", n_poli)

    if not slots:
        return {"feasible": True, "assignments": {}, "unassigned": [],
                "returned": returned, "poli": poli_added,
                "stats": {"total": 0, "assigned": 0, "unassigned": 0, "returned": len(returned)}}

    staff_list = [
        Staff(id=doc.id, name=doc.id, role="dokter",
              skills=frozenset(set(doc.activityIds) | {ok_act, pbk_act, poli_act}),
              carry_in=int(req.priorTotals.get(doc.id, 0)))   # regel 3: jaar-saldo als startstand
        for doc in doctors
    ]
    # let op: elke arts mag OK/PBK/Poli draaien (alle drie zijn arts-activiteiten);
    # koppeling beperkt alleen wie OK/PBK uit Excel mag, niet de poli-vulling.

    avail: dict[tuple[str, str], Avail] = {}
    for doc in doctors:
        mo_dates = {mo.get("date") for mo in doc.manualOff}
        for date_str in {s.date for s in slots}:
            wd = WD_CODES[date.fromisoformat(date_str).weekday()]
            if wd in doc.fixedOff or _is_biweekly_off(doc.biweeklyOff, date_str, wd) or date_str in mo_dates:
                avail[(doc.id, date_str)] = Avail.MANDATORY_OFF
            elif wd in doc.preferOff:
                avail.setdefault((doc.id, date_str), Avail.PREFER_OFF)
        for ab in doc.absences:
            f, t = ab.get("from"), ab.get("to")
            if not (f and t):
                continue
            kind = Avail.COURSE if ab.get("type") == "cursus" else Avail.VACATION
            for date_str in {s.date for s in slots}:
                if f <= date_str <= t:
                    avail[(doc.id, date_str)] = kind

    weights = SoftWeights(
        fairness=10, continuity=0, prefer_off=1,
        same_day_pair=15,  # regel 1: OK AM+PM zelfde persoon (> fairness, dwingt koppeling)
        poli_split=4,      # regel 2: niet 2× Poli op één dag (zacht, wijkt voor fairness/regel 1)
        ok_fairness=12,    # binnen-week OK-verdeling (> jaar-fairness, < koppeling)
        ok_skill=ok_act,
    )
    engine  = RosterEngine(staff_list, slots, avail, weights, partial_coverage=True,
                           pair_skill=ok_act, split_skill=poli_act)
    res     = engine.solve(max_seconds=30.0)

    assignments: dict[str, str] = {}
    unassigned:  list[str]      = []
    for slot in slots:
        who = res.assignments.get(slot.id, [])
        if who:
            assignments[slot.id] = who[0]
        else:
            unassigned.append(slot.id)

    return {
        "feasible":    len(unassigned) == 0,
        "assignments": assignments,    # "<date>__<period>__<type>__<i>" -> staffId
        "unassigned":  unassigned,
        "returned":    returned,       # [{date, period, type}]
        "poli":        poli_added,     # {date: {AM:n, PM:n}}
        "stats": {
            "total": len(slots), "assigned": len(assignments),
            "unassigned": len(unassigned), "returned": len(returned),
        },
    }


@app.post("/parse-requirements")
async def parse_requirements(
    ok_file:  UploadFile = File(None),
    pbk_file: UploadFile = File(None),
):
    result: dict = {}
    # alleen huidige + toekomstige weken: ondergrens = maandag van deze week
    today_d  = date.today()
    min_date = (today_d - timedelta(days=today_d.weekday())).isoformat()
    if ok_file  and ok_file.filename:
        _extract(await ok_file.read(),  "OK",  result, min_date)
    if pbk_file and pbk_file.filename:
        _extract(await pbk_file.read(), "PBK", result, min_date)

    count = sum(
        len(v.get("AM", [])) + len(v.get("PM", []))
        for v in result.values()
    )
    return {"requirements": result, "count": count}


@app.post("/solve-weekday")
def solve_weekday(req: SolveReq):
    start = date.fromisoformat(req.start)
    start = start - timedelta(days=start.weekday())   # uitlijnen op maandag

    # slots: ma-vr voor elke week
    slots = []
    for w in range(req.weeks):
        monday = start + timedelta(weeks=w)
        for i in range(5):
            d = monday + timedelta(days=i)
            slots.append(Slot(
                id=f"D{d.isoformat()}", date=d.isoformat(),
                period=Period.FULL_DAY, required_skill=SKILL,
                demand=1, week=w, seq=i,
            ))

    # staff + beschikbaarheid afleiden uit vaste vrije dagen / voorkeur / absenties
    staff, avail = [], {}
    for doc in req.doctors:
        staff.append(Staff(id=doc.id, name=doc.id, role="dokter",
                           skills=frozenset({SKILL}), carry_in=doc.carryIn))
        for slot in slots:
            wd = WD_CODES[date.fromisoformat(slot.date).weekday()]
            if wd in doc.fixedOff or _is_biweekly_off(doc.biweeklyOff, slot.date, wd):
                avail[(doc.id, slot.date)] = Avail.MANDATORY_OFF
            elif wd in doc.preferOff:
                avail.setdefault((doc.id, slot.date), Avail.PREFER_OFF)
        for ab in doc.absences:
            f, t = ab.get("from"), ab.get("to")
            if not f or not t:
                continue
            kind = Avail.COURSE if ab.get("type") == "cursus" else Avail.VACATION
            for slot in slots:
                if f <= slot.date <= t:
                    avail[(doc.id, slot.date)] = kind

    w = req.weights or {}
    weights = SoftWeights(
        fairness=w.get("fairness", 10),
        continuity=w.get("continuity", 3),
        prefer_off=w.get("preferOff", 1),
    )

    res = RosterEngine(staff, slots, avail, weights).solve()
    if not res.feasible:
        return {"feasible": False, "status": res.status, "assignments": {}, "totals": {}, "objective": 0}

    assignments = {}
    for slot in slots:
        who = res.assignments.get(slot.id, [])
        if who:
            assignments[slot.date] = who[0]

    return {
        "feasible": True, "status": res.status,
        "assignments": assignments, "totals": res.totals, "objective": res.objective,
    }
